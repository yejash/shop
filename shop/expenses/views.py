from django.shortcuts import render

# Create your views here.
import json
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponseBadRequest
from django.contrib.auth.decorators import login_required
from .models import Expense
from django.views.decorators.http import require_http_methods
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from django.utils.dateparse import parse_date
from decimal import Decimal, InvalidOperation
from django.utils.timezone import now


@login_required
def expenses_page(request):
    # Render page (template will fetch list via AJAX)
    return render(request, 'expenses/expenses.html')

@login_required
def api_list_expenses(request):
    expenses = Expense.objects.filter(user=request.user).order_by('-date', '-created_at')
    data = [e.as_dict() for e in expenses]
    return JsonResponse(data, safe=False)

@login_required
@require_http_methods(["POST"])
def api_add_expense(request):
    try:
        payload = json.loads(request.body)
        e = Expense.objects.create(
            user=request.user,
            date=payload.get('date'),
            description=payload.get('description'),
            mode=payload.get('mode'),
            amount=payload.get('amount')
        )
        return JsonResponse({'status': 'success', 'id': e.id})
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

@login_required
@require_http_methods(["POST"])
def api_update_expense(request):
    try:
        payload = json.loads(request.body)
        eid = payload.get('id')
        e = get_object_or_404(Expense, id=eid, user=request.user)
        e.date = payload.get('date')
        e.description = payload.get('description')
        e.mode = payload.get('mode')
        e.amount = payload.get('amount')
        e.save()
        return JsonResponse({'status': 'success'})
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

@login_required
@require_http_methods(["POST"])
def api_delete_expense(request):
    try:
        payload = json.loads(request.body)
        eid = payload.get('id')
        e = get_object_or_404(Expense, id=eid, user=request.user)
        e.delete()
        return JsonResponse({'status': 'success'})
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

def api_add(request):
    if request.method == "POST":
        data = json.loads(request.body)
        expense = Expense.objects.create(
            date=data["date"],
            description=data["description"],
            mode=data["mode"],
            amount=data["amount"],
            user=request.user  # if linked to logged-in user
        )
        return JsonResponse({"status": "success", "id": expense.id})
    return JsonResponse({"status": "error"}, status=400)

# ---------- Excel Export ----------
@login_required
def export_expenses_excel(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    qs = Expense.objects.filter(user=request.user).order_by("date")

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = ["Date", "Description", "Mode", "Amount (₹)"]
    ws.append(headers)

    total = 0
    for e in qs:
        ws.append([e.date.strftime("%d-%b-%Y"), e.description, e.mode, float(e.amount)])
        total += float(e.amount)

    # Total row
    ws.append(["", "", "Total", total])

    # Auto column width
    for col_idx, _ in enumerate(headers, start=1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell_val = row[0].value
            max_len = max(max_len, len(str(cell_val)) if cell_val else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"expenses_{now().date().isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ---------- PDF Export ----------
@login_required
def export_expenses_pdf(request):
    from reportlab.lib.pagesizes import A4, inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    qs = Expense.objects.filter(user=request.user).order_by("date")

    # response + doc
    response = HttpResponse(content_type="application/pdf")
    filename = f"expenses_{now().date().isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph("Expenses Report", styles["Title"]))
    story.append(Spacer(1, 0.2 * inch))

    data = [["Date", "Description", "Mode", "Amount (₹)"]]
    total = 0
    for e in qs:
        data.append([e.date.strftime("%d-%b-%Y"), e.description, e.mode, f"{float(e.amount):.2f}"])
        total += float(e.amount)

    # Total row
    data.append(["", "", "Total", f"{total:.2f}"])

    table = Table(data, colWidths=[1.3*inch, 3.2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (3,1), (3,-1), "RIGHT"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.whitesmoke, colors.beige]),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
    ]))

    story.append(table)
    doc.build(story)
    return response