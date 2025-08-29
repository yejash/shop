import json
from datetime import date
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponseBadRequest
from django.shortcuts import render, get_object_or_404
from .models import Income, MODE_CHOICES
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from django.utils.dateparse import parse_date
from decimal import Decimal, InvalidOperation
from django.utils.timezone import now



def _allowed_modes():
    return [m[0] for m in MODE_CHOICES]

@login_required
def income_list(request):
    # Render the page; data is fetched via AJAX
    return render(request, 'income/income_list.html', {'modes': MODE_CHOICES})

@login_required
def api_list(request):
    if request.method != 'GET':
        return HttpResponseBadRequest('GET only')
    qs = Income.objects.filter(user=request.user).order_by('-date', '-id')
    data = [
        {
            'id': obj.id,
            'date': obj.date.isoformat(),
            'description': obj.description,
            'mode': obj.mode,
            'amount': float(obj.amount),
        }
        for obj in qs
    ]
    return JsonResponse(data, safe=False)


@login_required
def api_add(request):
    if request.method != "POST":
        return HttpResponseBadRequest("POST only")
    try:
        payload = json.loads(request.body.decode('utf-8') or "{}")
        date_str = payload.get('date')
        description = (payload.get('description') or "").strip()
        mode = payload.get('mode')
        amount_raw = payload.get('amount')

        if not (date_str and description and mode and amount_raw is not None):
            return JsonResponse({'status': 'error', 'message': 'Missing fields'}, status=400)

        d = parse_date(date_str)
        if not d:
            return JsonResponse({'status': 'error', 'message': 'Invalid date'}, status=400)

        try:
            amount = Decimal(str(amount_raw))
        except InvalidOperation:
            return JsonResponse({'status': 'error', 'message': 'Invalid amount'}, status=400)

        # create model instance
        obj = Income.objects.create(
            user=request.user,
            date=d,
            description=description,
            mode=mode,
            amount=amount
        )

        return JsonResponse({'status':'success', 'id': obj.id})
    except Exception as e:
        # for dev only: include exception message to help debugging
        return JsonResponse({'status':'error','message': str(e)}, status=500)


@login_required
def api_update(request):
    if request.method != 'POST':
        return HttpResponseBadRequest('POST only')
    try:
        payload = json.loads(request.body.decode('utf-8'))
        pk = payload.get('id')
        if not pk:
            return JsonResponse({'status': 'error', 'message': 'Missing id'}, status=400)
        obj = get_object_or_404(Income, pk=pk, user=request.user)

        if 'date' in payload and payload['date']:
            obj.date = date.fromisoformat(payload['date'])
        if 'description' in payload:
            obj.description = (payload['description'] or '').strip()
        if 'mode' in payload:
            if payload['mode'] not in _allowed_modes():
                return JsonResponse({'status': 'error', 'message': 'Invalid mode'}, status=400)
            obj.mode = payload['mode']
        if 'amount' in payload and payload['amount'] is not None:
            obj.amount = payload['amount']
        obj.save()

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required
def api_delete(request):
    if request.method != 'POST':
        return HttpResponseBadRequest('POST only')
    try:
        payload = json.loads(request.body.decode('utf-8'))
        pk = payload.get('id')
        obj = get_object_or_404(Income, pk=pk, user=request.user)
        obj.delete()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# ---------- Excel ----------
@login_required
def export_income_excel(request):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    qs = Income.objects.filter(user=request.user).order_by("date")

    wb = Workbook()
    ws = wb.active
    ws.title = "Income"

    headers = ["Date", "Description", "Mode", "Amount (₹)"]
    ws.append(headers)

    total = 0
    for i in qs:
        ws.append([i.date.strftime("%d-%b-%Y"), i.description, i.mode, float(i.amount)])
        total += float(i.amount)

    ws.append(["", "", "Total", total])

    # simple auto width
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            v = row[0].value
            max_len = max(max_len, len(str(v)) if v else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="income_{now().date().isoformat()}.xlsx"'
    wb.save(resp)
    return resp


# ---------- PDF ----------
@login_required
def export_income_pdf(request):
    from reportlab.lib.pagesizes import A4, inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    qs = Income.objects.filter(user=request.user).order_by("date")

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="income_{now().date().isoformat()}.pdf"'

    doc = SimpleDocTemplate(resp, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Income Report", styles["Title"]))
    story.append(Spacer(1, 0.2 * inch))

    data = [["Date", "Description", "Mode", "Amount (₹)"]]
    total = 0
    for i in qs:
        data.append([i.date.strftime("%d-%b-%Y"), i.description, i.mode, f"{float(i.amount):.2f}"])
        total += float(i.amount)
    data.append(["", "", "Total", f"{total:.2f}"])

    table = Table(data, colWidths=[1.3*inch, 3.2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (3,1), (3,-1), "RIGHT"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.whitesmoke, colors.beige]),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
    ]))

    story.append(table)
    doc.build(story)
    return resp